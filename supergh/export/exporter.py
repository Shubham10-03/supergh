"""Export system — CSV, JSON, Markdown, Excel export for any command output."""

from __future__ import annotations

import csv
import json
from io import StringIO
from pathlib import Path
from typing import Any

from rich.console import Console

console = Console()


def export_data(data: list[dict], path: str) -> None:
    """Export a list of dicts to file based on extension."""
    p = Path(path)
    ext = p.suffix.lower()

    if ext == ".json":
        _export_json(data, p)
    elif ext == ".csv":
        _export_csv(data, p)
    elif ext == ".md":
        _export_markdown(data, p)
    elif ext in (".xlsx", ".xls"):
        _export_excel(data, p)
    else:
        # Default to JSON
        _export_json(data, p)

    console.print(f"[green]Exported {len(data)} rows to {p}[/green]")


def _export_json(data: list[dict], path: Path) -> None:
    path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def _export_csv(data: list[dict], path: Path) -> None:
    if not data:
        path.write_text("", encoding="utf-8")
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)


def _export_markdown(data: list[dict], path: Path) -> None:
    if not data:
        path.write_text("", encoding="utf-8")
        return
    headers = list(data[0].keys())
    lines = []
    # Header row
    lines.append("| " + " | ".join(headers) + " |")
    # Separator
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    # Data rows
    for row in data:
        values = [str(row.get(h, "")).replace("|", "\\|").replace("\n", " ")[:80] for h in headers]
        lines.append("| " + " | ".join(values) + " |")
    path.write_text("\n".join(lines), encoding="utf-8")


def _export_excel(data: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        console.print("[red]Excel export requires openpyxl. Install with: pip install openpyxl[/red]")
        raise SystemExit(1)

    if not data:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    headers = list(data[0].keys())

    # Header row with styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(path)


def format_for_output(data: list[dict], output_format: str) -> str:
    """Format data as string for stdout (json or csv)."""
    if output_format == "json":
        return json.dumps(data, indent=2, default=str)
    elif output_format == "csv":
        if not data:
            return ""
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()
    return json.dumps(data, indent=2, default=str)
